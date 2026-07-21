"""
Carga de documentos y división en chunks.

Soporta PDF (PyMuPDFLoader) y Excel (openpyxl, fila por fila).
TODO: sumar loaders para Word, PowerPoint, Markdown, CSV, JSON y HTML
      a medida que el proyecto lo requiera (ver el desafío para la lista completa).
"""
from pathlib import Path

import openpyxl
from langchain_core.documents import Document
from langchain_community.document_loaders import PyMuPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter

from app.config import DOCUMENTS_PATH


def _cargar_excel(ruta: Path) -> list:
    """Convierte cada fila de cada hoja en un Document independiente."""
    documentos = []
    wb = openpyxl.load_workbook(ruta, data_only=True)

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        filas = list(hoja.iter_rows(values_only=True))

        if not filas:
            continue

        encabezados = [str(h) if h is not None else "" for h in filas[0]]

        for numero_fila, fila in enumerate(filas[1:], start=2):
            partes = []
            for encabezado, valor in zip(encabezados, fila):
                if valor is None or valor == "":
                    continue
                partes.append(f"{encabezado}: {valor}")

            if not partes:
                continue

            texto = " | ".join(partes)

            documentos.append(
                Document(
                    page_content=texto,
                    metadata={
                        "source": str(ruta),
                        "sheet": nombre_hoja,
                        "row": numero_fila,
                    },
                )
            )

    return documentos


def cargar_documentos(carpeta: str = DOCUMENTS_PATH) -> list:
    docs = []
    carpeta_path = Path(carpeta)

    if not carpeta_path.exists():
        raise FileNotFoundError(
            f"La carpeta '{carpeta}' no existe. Créala y coloca ahí tus documentos."
        )

    for documento in carpeta_path.glob("*.pdf"):
        try:
            loader = PyMuPDFLoader(str(documento))
            docs.extend(loader.load())
            print(f"Archivo cargado: {documento.name}")
        except Exception as e:
            print(f"Error cargando archivo: {documento.name}: {e}")

    for documento in carpeta_path.glob("*.xlsx"):
        try:
            docs.extend(_cargar_excel(documento))
            print(f"Archivo cargado: {documento.name}")
        except Exception as e:
            print(f"Error cargando archivo: {documento.name}: {e}")

    print(f"Total de documentos cargados: {len(docs)}")
    return docs


def dividir_en_chunks(docs: list, chunk_size: int = 300, chunk_overlap: int = 30) -> list:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size, chunk_overlap=chunk_overlap
    )
    return splitter.split_documents(docs)
